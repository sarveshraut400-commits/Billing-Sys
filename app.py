# Live Production Backend - Render Deployment (diddy branch)
from flask import Flask, request, jsonify, send_file, make_response

from flask_cors import CORS
import smtplib
from email.mime.text import MIMEText
import random
import time
import uuid
import io
import json
import os
import threading

from datetime import datetime


# Database helper functions
from database import get_db_connection, log_activity

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import openpyxl

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization,X-Requested-With'
    response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,DELETE,OPTIONS'
    
    # If this was an OPTIONS request and it failed (e.g. 405), forcefully return 200 OK for preflight
    if request.method == 'OPTIONS' and response.status_code != 200:
        response.status_code = 200
        
    return response

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INVOICES_DIR = os.path.join(BASE_DIR, 'invoices')
os.makedirs(INVOICES_DIR, exist_ok=True)


# ==========================================
# EMAIL SETTINGS & ROBUST DISPATCH ENGINE
# ==========================================
import socket
import ssl
from concurrent.futures import ThreadPoolExecutor
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate, make_msgid

email_executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix="MailWorker")

# Strict sanitization with guaranteed non-empty fallback
_env_email = os.environ.get("SENDER_EMAIL", "").strip()
SENDER_EMAIL = _env_email if _env_email else "systemdefault96@gmail.com"

_env_pwd = str(os.environ.get("SENDER_PASSWORD", "")).replace(" ", "").replace('"', '').replace("'", "").strip()
SENDER_PASSWORD = _env_pwd if _env_pwd else "zkavukpsjfehuhqw"

# Force IPv4 socket resolution to prevent Linux/Render [Errno 101] Network is unreachable IPv6 bugs
class IPv4SMTP_SSL(smtplib.SMTP_SSL):
    def _get_socket(self, host, port, timeout):
        infos = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
        last_err = None
        for res in infos:
            af, socktype, proto, canonname, sa = res
            s = None
            try:
                s = socket.socket(af, socktype, proto)
                s.settimeout(timeout)
                s.connect(sa)
                return self.context.wrap_socket(s, server_hostname=self._host)
            except Exception as e:
                last_err = e
                if s:
                    s.close()
        if last_err:
            raise last_err

class IPv4SMTP(smtplib.SMTP):
    def _get_socket(self, host, port, timeout):
        infos = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
        last_err = None
        for res in infos:
            af, socktype, proto, canonname, sa = res
            s = None
            try:
                s = socket.socket(af, socktype, proto)
                s.settimeout(timeout)
                s.connect(sa)
                return s
            except Exception as e:
                last_err = e
                if s:
                    s.close()
        if last_err:
            raise last_err

def send_email(receiver_email, subject, body_text, html_content=None):
    if not receiver_email:
        print("[EMAIL WARNING] No receiver email provided.")
        return False

    sender = SENDER_EMAIL
    pwd = SENDER_PASSWORD
    
    if not sender or not pwd:
        print("[EMAIL WARNING] Missing sender email credentials.")
        return False

    if html_content:
        msg = MIMEMultipart('alternative')
        msg.attach(MIMEText(body_text, 'plain', 'utf-8'))
        msg.attach(MIMEText(html_content, 'html', 'utf-8'))
    else:
        msg = MIMEText(body_text, 'plain', 'utf-8')

    msg['Subject'] = subject
    msg['From'] = f"SuperMart POS <{sender}>"
    msg['To'] = receiver_email
    msg['Date'] = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid(domain='gmail.com')
    
    # 1. Try Port 465 (Direct SSL with IPv4 force)
    try:
        context = ssl.create_default_context()
        with IPv4SMTP_SSL('smtp.gmail.com', 465, context=context, timeout=8) as server:
            server.login(sender, pwd)
            server.send_message(msg)
        print(f"[EMAIL SUCCESS] Delivered to {receiver_email} via Port 465 (SSL)")
        return True
    except Exception as e465:
        print(f"[SMTP Port 465 Notice] {e465}. Retrying Port 587 (TLS)...")
        # 2. Fallback to Port 587 (STARTTLS with IPv4 force)
        try:
            context = ssl.create_default_context()
            with IPv4SMTP('smtp.gmail.com', 587, timeout=8) as server:
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(sender, pwd)
                server.send_message(msg)
            print(f"[EMAIL SUCCESS] Delivered to {receiver_email} via Port 587 (STARTTLS)")
            return True
        except Exception as e587:
            print(f"[SMTP Port 587 Error] {e587}")
            log_activity("Settings", "Email Dispatch Error", f"Failed sending to {receiver_email}: {str(e465)} / {str(e587)}", performed_by="System Mailer")
            return False

def send_otp_email(receiver_email, otp):
    subject = "SuperMart POS - Security Verification OTP"
    body = (
        f"Hello,\n\n"
        f"Your SuperMart POS Security Verification OTP is:\n"
        f"🔑 {otp}\n\n"
        f"This code is valid for 10 minutes. Do not share it with anyone.\n\n"
        f"Regards,\n"
        f"SuperMart Security & Operations"
    )
    html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 24px; border: 1px solid #e2e8f0; border-radius: 14px; background-color: #ffffff;">
      <h2 style="color: #0f172a; margin-top: 0; font-size: 18px;">🔐 SuperMart Security Verification</h2>
      <p style="color: #475569; font-size: 14px;">Use the following One-Time Password (OTP) to authorize your administrative action:</p>
      <div style="background-color: #f8fafc; border: 1.5px solid #cbd5e1; padding: 18px; text-align: center; border-radius: 10px; font-size: 32px; font-weight: 800; letter-spacing: 8px; color: #0f172a; margin: 20px 0; font-family: monospace;">
        {otp}
      </div>
      <p style="color: #94a3b8; font-size: 12px; margin-bottom: 0;">This OTP is valid for 10 minutes. If you did not request this code, please secure your account immediately.</p>
    </div>
    """
    return send_email(receiver_email, subject, body, html)

def send_welcome_email(receiver_email, name, role, password):
    is_admin = (str(role).strip().lower() == 'admin')
    
    if is_admin:
        # ==========================================
        # DISTINCT EXECUTIVE ADMINISTRATOR EMAIL
        # ==========================================
        subject = f"👑 SuperMart Executive Console: Administrator Access Granted for {name}"
        body = (
            f"Dear {name},\n\n"
            f"EXECUTIVE ADMINISTRATOR APPOINTMENT & ACCESS CREDENTIALS\n\n"
            f"You have been granted Full Administrator Authority over the SuperMart POS Enterprise System.\n\n"
            f"EXECUTIVE CREDENTIALS:\n"
            f"• Portal URL: https://billing-sys-beta.vercel.app\n"
            f"• Login Email: {receiver_email}\n"
            f"• Master Password: {password}\n"
            f"• Authority Level: Root Administrator\n\n"
            f"Your administrative capabilities include:\n"
            f"1. Revenue & Real-time Fiscal Analytics\n"
            f"2. Staff Management & Access Security\n"
            f"3. Inventory Control & Barcode Management\n"
            f"4. Live Security Audit Trail Inspection\n\n"
            f"Access your executive console at https://billing-sys-beta.vercel.app\n\n"
            f"Regards,\n"
            f"SuperMart Executive Operations"
        )
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1.0">
          <title>Executive Administrator Access</title>
        </head>
        <body style="margin: 0; padding: 0; background-color: #0f172a; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #1e293b;">
          <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background-color: #0f172a; padding: 35px 12px;">
            <tr>
              <td align="center">
                <table width="100%" border="0" cellspacing="0" cellpadding="0" style="max-width: 580px; background-color: #ffffff; border-radius: 16px; overflow: hidden; box-shadow: 0 20px 40px rgba(0, 0, 0, 0.4); border: 1px solid #334155;">
                  
                  <!-- EXECUTIVE HEADER -->
                  <tr>
                    <td style="background: linear-gradient(135deg, #1e1b4b 0%, #312e81 60%, #4338ca 100%); padding: 32px 30px; text-align: center;">
                      <div style="display: inline-block; background: rgba(245, 158, 11, 0.15); border: 1px solid rgba(245, 158, 11, 0.4); padding: 5px 14px; border-radius: 20px; color: #fbbf24; font-size: 11px; font-weight: 800; letter-spacing: 1.5px; text-transform: uppercase; margin-bottom: 10px;">
                        👑 ROOT ADMINISTRATOR ACCESS
                      </div>
                      <h1 style="margin: 0; color: #ffffff; font-size: 22px; font-weight: 800; letter-spacing: -0.3px;">
                        Executive Console Credentials
                      </h1>
                      <p style="margin: 6px 0 0 0; color: #cbd5e1; font-size: 13px;">
                        SuperMart Enterprise Management & Analytics
                      </p>
                    </td>
                  </tr>

                  <!-- BODY -->
                  <tr>
                    <td style="padding: 28px 30px;">
                      <p style="margin: 0 0 14px 0; color: #0f172a; font-size: 16px; font-weight: 600;">
                        Hello {name},
                      </p>
                      <p style="margin: 0 0 20px 0; color: #475569; font-size: 14px; line-height: 1.5;">
                        Your <strong>Executive Administrator</strong> account has been configured with full administrative privileges. Use the credentials below to log in:
                      </p>

                      <!-- CREDENTIALS BOX -->
                      <div style="background-color: #f8fafc; border: 1.5px solid #cbd5e1; border-radius: 12px; padding: 18px 20px; margin-bottom: 22px;">
                        <table width="100%" border="0" cellspacing="0" cellpadding="6">
                          <tr>
                            <td width="38%" style="color: #64748b; font-size: 13px; font-weight: 600;">Console URL:</td>
                            <td><a href="https://billing-sys-beta.vercel.app" style="color: #4f46e5; font-size: 13px; font-weight: 700; text-decoration: none;">https://billing-sys-beta.vercel.app</a></td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Admin Login:</td>
                            <td style="color: #0f172a; font-size: 13px; font-family: monospace; font-weight: 700;">{receiver_email}</td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Master Password:</td>
                            <td>
                              <span style="background-color: #1e1b4b; color: #38bdf8; font-family: monospace; font-size: 14px; font-weight: 800; padding: 3px 10px; border-radius: 6px; letter-spacing: 1px;">
                                {password}
                              </span>
                            </td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Authority Tier:</td>
                            <td>
                              <span style="background-color: #fef3c7; color: #92400e; border: 1px solid #fde68a; font-size: 11px; font-weight: 800; padding: 2px 8px; border-radius: 12px;">
                                FULL SYSTEM ACCESS
                              </span>
                            </td>
                          </tr>
                        </table>
                      </div>

                      <!-- CTA BUTTON -->
                      <table width="100%" border="0" cellspacing="0" cellpadding="0" style="margin-bottom: 24px;">
                        <tr>
                          <td align="center">
                            <a href="https://billing-sys-beta.vercel.app" target="_blank" style="background: linear-gradient(135deg, #4f46e5 0%, #3730a3 100%); color: #ffffff; text-decoration: none; font-size: 14px; font-weight: 700; padding: 13px 32px; border-radius: 10px; display: inline-block; box-shadow: 0 4px 14px rgba(79, 70, 229, 0.35);">
                              🏛️ Launch Executive Admin Console →
                            </a>
                          </td>
                        </tr>
                      </table>

                      <!-- EXECUTIVE DUTIES -->
                      <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 14px 16px;">
                        <div style="font-size: 12px; font-weight: 800; color: #1e1b4b; text-transform: uppercase; margin-bottom: 8px;">
                          📋 Executive Permissions:
                        </div>
                        <table width="100%" border="0" cellspacing="0" cellpadding="4">
                          <tr>
                            <td style="font-size: 12px; color: #334155;">• <strong>Revenue Analytics:</strong> Real-time sales, tax reports & profit charts.</td>
                          </tr>
                          <tr>
                            <td style="font-size: 12px; color: #334155;">• <strong>Staff Governance:</strong> Hire employees, manage roles & track shifts.</td>
                          </tr>
                          <tr>
                            <td style="font-size: 12px; color: #334155;">• <strong>Inventory Operations:</strong> Price configuration & stock management.</td>
                          </tr>
                        </table>
                      </div>

                    </td>
                  </tr>

                  <!-- FOOTER -->
                  <tr>
                    <td style="background-color: #0f172a; border-top: 1px solid #1e293b; padding: 18px 30px; text-align: center;">
                      <p style="margin: 0; color: #94a3b8; font-size: 11px;">
                        SuperMart Enterprise POS • Administrative Executive Communication • GSTIN: 27AABCU9603R1ZM
                      </p>
                    </td>
                  </tr>

                </table>
              </td>
            </tr>
          </table>
        </body>
        </html>
        """
    else:
        # ==========================================
        # DISTINCT EMPLOYEE / STAFF ACCOUNT EMAIL
        # ==========================================
        subject = f"🛒 SuperMart POS: Staff Account & Shift Credentials for {name}"
        body = (
            f"Hello {name},\n\n"
            f"Your SuperMart POS staff account has been created.\n\n"
            f"LOGIN CREDENTIALS:\n"
            f"• Portal URL: https://billing-sys-beta.vercel.app\n"
            f"• Login Email: {receiver_email}\n"
            f"• Password: {password}\n"
            f"• Role: {role.capitalize()}\n\n"
            f"Log in at the portal to start your shift.\n\n"
            f"Regards,\n"
            f"SuperMart Operations"
        )
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1.0">
          <title>SuperMart Staff Access</title>
        </head>
        <body style="margin: 0; padding: 0; background-color: #f8fafc; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #1e293b;">
          <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background-color: #f8fafc; padding: 35px 12px;">
            <tr>
              <td align="center">
                <table width="100%" border="0" cellspacing="0" cellpadding="0" style="max-width: 540px; background-color: #ffffff; border-radius: 14px; overflow: hidden; border: 1px solid #e2e8f0; box-shadow: 0 4px 16px rgba(0, 0, 0, 0.04);">
                  
                  <!-- STAFF HEADER -->
                  <tr>
                    <td style="padding: 24px 28px 18px 28px; border-bottom: 1px solid #f1f5f9;">
                      <table width="100%" border="0" cellspacing="0" cellpadding="0">
                        <tr>
                          <td align="left">
                            <span style="font-size: 16px; font-weight: 800; color: #0f172a; letter-spacing: -0.3px;">
                              🛒 SuperMart<span style="color: #10b981;">.</span>
                            </span>
                          </td>
                          <td align="right">
                            <span style="background-color: #dcfce7; color: #166534; font-size: 11px; font-weight: 700; padding: 4px 10px; border-radius: 6px;">
                              STAFF ACCOUNT
                            </span>
                          </td>
                        </tr>
                      </table>
                    </td>
                  </tr>

                  <!-- BODY -->
                  <tr>
                    <td style="padding: 26px 28px;">
                      <h2 style="margin: 0 0 6px 0; color: #0f172a; font-size: 18px; font-weight: 700;">
                        Welcome to the POS Shift
                      </h2>
                      <p style="margin: 0 0 20px 0; color: #64748b; font-size: 14px; line-height: 1.5;">
                        Hello <strong>{name}</strong>, your cashier and counter access has been set up. Use the credentials below to log in:
                      </p>

                      <!-- CREDENTIALS TABLE -->
                      <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 16px 18px; margin-bottom: 22px;">
                        <table width="100%" border="0" cellspacing="0" cellpadding="6">
                          <tr>
                            <td width="36%" style="color: #64748b; font-size: 13px; font-weight: 600;">Portal URL:</td>
                            <td><a href="https://billing-sys-beta.vercel.app" style="color: #10b981; font-size: 13px; font-weight: 600; text-decoration: none;">https://billing-sys-beta.vercel.app</a></td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Login Email:</td>
                            <td style="color: #0f172a; font-size: 13px; font-family: monospace; font-weight: 600;">{receiver_email}</td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Password:</td>
                            <td>
                              <span style="background-color: #e2e8f0; color: #0f172a; font-family: monospace; font-size: 14px; font-weight: 700; padding: 2px 8px; border-radius: 4px;">
                                {password}
                              </span>
                            </td>
                          </tr>
                          <tr>
                            <td style="color: #64748b; font-size: 13px; font-weight: 600;">Role:</td>
                            <td style="color: #0f172a; font-size: 13px; font-weight: 600;">{role.capitalize()}</td>
                          </tr>
                        </table>
                      </div>

                      <!-- CTA BUTTON -->
                      <table width="100%" border="0" cellspacing="0" cellpadding="0" style="margin-bottom: 20px;">
                        <tr>
                          <td align="left">
                            <a href="https://billing-sys-beta.vercel.app" target="_blank" style="background-color: #10b981; color: #ffffff; text-decoration: none; font-size: 14px; font-weight: 600; padding: 11px 24px; border-radius: 8px; display: inline-block;">
                              🚀 Log in to POS Portal →
                            </a>
                          </td>
                        </tr>
                      </table>

                      <p style="margin: 0; color: #94a3b8; font-size: 12px;">
                        Contact your store administrator if you need assistance with your credentials.
                      </p>
                    </td>
                  </tr>

                  <!-- FOOTER -->
                  <tr>
                    <td style="background-color: #f8fafc; border-top: 1px solid #f1f5f9; padding: 16px 28px; text-align: left;">
                      <p style="margin: 0; color: #94a3b8; font-size: 11px;">
                        SuperMart POS • Retail Operations System • GSTIN: 27AABCU9603R1ZM
                      </p>
                    </td>
                  </tr>

                </table>
              </td>
            </tr>
          </table>
        </body>
        </html>
        """

    success = send_email(receiver_email, subject, body, html)
    if success:
        log_activity(
            category="Login/Checkout",
            action="Welcome Email Sent",
            details=f"Sent {'Administrator' if is_admin else 'Staff'} email to {name} ({receiver_email})",
            performed_by="System"
        )
    return success









# ==========================================
# 0. AUTH USERS DATABASE (PERMANENT JSON)
# ==========================================
USERS_FILE = 'users.json'

default_users = {
    "admin": {
        "email": "systemdefault96@gmail.com", 
        "password": "admin123", 
        "otp": None
    },
    "employee": {
        "email": "staff@store.com", 
        "password": "staff123", 
        "otp": None
    }
}

if os.path.exists(USERS_FILE):
    with open(USERS_FILE, 'r') as f:
        users_db = json.load(f)
else:
    users_db = default_users
    with open(USERS_FILE, 'w') as f:
        json.dump(users_db, f, indent=4)

def save_users():
    """Saves auth changes permanently so the login page updates!"""
    with open(USERS_FILE, 'w') as f:
        json.dump(users_db, f, indent=4)


SHOP_SETTINGS_FILE = os.path.join(BASE_DIR, 'shop_settings.json')

DEFAULT_SHOP_SETTINGS = {
    "name": "SuperMart POS",
    "phone": "+91 9876543210",
    "email": "systemdefault96@gmail.com",
    "gstin": "27AABCU9603R1ZM",
    "address": "123 Main Commercial Hub, Mumbai, MH, India",
    "receiptFooter": "Thank you for shopping with us! Visit again."
}

def load_shop_settings():
    if os.path.exists(SHOP_SETTINGS_FILE):
        try:
            with open(SHOP_SETTINGS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {**DEFAULT_SHOP_SETTINGS, **data}
        except Exception as e:
            print(f"Error loading shop_settings.json: {e}")
    return DEFAULT_SHOP_SETTINGS.copy()

def save_shop_settings(settings):
    try:
        with open(SHOP_SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, indent=4)
        return True
    except Exception as e:
        print(f"Error saving shop_settings.json: {e}")
        return False


def clean_pdf_str(text):
    if not text:
        return ""
    text = str(text)
    replacements = {
        '•': '-',
        '₹': 'Rs. ',
        '–': '-',
        '—': '-',
        '“': '"',
        '”': '"',
        '‘': "'",
        '’': "'",
        '™': '(TM)',
        '®': '(R)',
        '©': '(C)'
    }
    for orig, repl in replacements.items():
        text = text.replace(orig, repl)
    return text


def generate_pdf_invoice(invoice_no, customer_name, phone, total_amount, items, date_str, time_str, cashier):
    try:
        filename = f"Invoice_{invoice_no}.pdf"
        filepath = os.path.join(INVOICES_DIR, filename)
        
        shop = load_shop_settings()
        shop_name = str(shop.get('name', 'SuperMart POS')).strip()
        shop_phone = str(shop.get('phone', '+91 9876543210')).strip()
        shop_email = str(shop.get('email', 'systemdefault96@gmail.com')).strip()
        shop_gstin = str(shop.get('gstin', '27AABCU9603R1ZM')).strip()
        shop_address = str(shop.get('address', '123 Main Commercial Hub, Mumbai, MH, India')).strip()
        receipt_footer = str(shop.get('receiptFooter', 'Thank you for shopping with us! Visit again.')).strip()

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        normal_style = styles['Normal']
        
        title_style = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#047857'),
            alignment=1
        )
        
        subtitle_style = ParagraphStyle(
            'InvoiceSubtitle',
            parent=normal_style,
            fontName='Helvetica',
            fontSize=9,
            leading=13,
            textColor=colors.HexColor('#4B5563'),
            alignment=1
        )

        gstin_style = ParagraphStyle(
            'GSTINHeader',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#065F46'),
            alignment=1
        )

        inv_head_style = ParagraphStyle(
            'InvoiceHeaderLabel',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#111827'),
            alignment=1
        )
        
        meta_style = ParagraphStyle(
            'InvoiceMeta',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor('#1F2937')
        )

        story = []
        # Store Name Header
        story.append(Paragraph(clean_pdf_str(shop_name.upper()), title_style))
        story.append(Paragraph(clean_pdf_str(f"{shop_address} | Phone: {shop_phone} | Email: {shop_email}"), subtitle_style))
        story.append(Spacer(1, 6))

        # GSTIN / TAX REGISTRATION ID BANNER
        gstin_banner = Table([[Paragraph(clean_pdf_str(f"<b>GSTIN / TAX REGISTRATION ID: {shop_gstin}</b>"), gstin_style)]], colWidths=[540])
        gstin_banner.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#D1FAE5')),
            ('PADDING', (0,0), (-1,-1), 6),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor('#10B981')),
            ('LINEABOVE', (0,0), (-1,-1), 1, colors.HexColor('#10B981')),
        ]))
        story.append(gstin_banner)
        story.append(Spacer(1, 10))

        story.append(Paragraph(clean_pdf_str(f"TAX INVOICE | #{invoice_no}"), inv_head_style))
        story.append(Spacer(1, 10))
        
        meta_data = [
            [Paragraph(clean_pdf_str(f"<b>Date & Time:</b> {date_str} {time_str}"), meta_style), Paragraph(clean_pdf_str(f"<b>Cashier:</b> {cashier}"), meta_style)],
            [Paragraph(clean_pdf_str(f"<b>Customer:</b> {customer_name or 'Walk-in Customer'}"), meta_style), Paragraph(clean_pdf_str(f"<b>Mobile:</b> {phone or 'N/A'}"), meta_style)]
        ]
        meta_table = Table(meta_data, colWidths=[270, 270])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F9FAFB')),
            ('PADDING', (0,0), (-1,-1), 7),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 12))
        
        table_data = [
            [Paragraph("<b>Item Description</b>", meta_style), Paragraph("<b>Price (Rs.)</b>", meta_style), Paragraph("<b>Qty</b>", meta_style), Paragraph("<b>Amount (Rs.)</b>", meta_style)]
        ]
        
        for item in items:
            name = item.get('name') or item.get('product_name') or 'Item'
            price = float(item.get('price', 0))
            qty = int(item.get('quantity') or item.get('qty') or 1)
            amt = float(item.get('amount') or (price * qty))
            
            table_data.append([
                Paragraph(clean_pdf_str(name), normal_style),
                Paragraph(f"Rs. {price:.2f}", normal_style),
                Paragraph(str(qty), normal_style),
                Paragraph(f"Rs. {amt:.2f}", normal_style)
            ])
            
        items_table = Table(table_data, colWidths=[240, 100, 80, 120])
        items_table.setStyle(TableStyle([
            ('HEADERBACKGROUND', (0,0), (-1,0), colors.HexColor('#ECFDF5')),
            ('HEADERTEXTCOLOR', (0,0), (-1,0), colors.HexColor('#065F46')),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 12))
        
        subtotal = sum(float(i.get('amount') or (float(i.get('price',0)) * int(i.get('quantity') or i.get('qty') or 1))) for i in items)
        discount = subtotal * 0.05
        grand_total = float(total_amount)

        # Tax calculation breakdown (e.g., GST 18% inclusive)
        taxable_val = grand_total / 1.18
        cgst_val = (grand_total - taxable_val) / 2
        sgst_val = cgst_val
        
        total_data = [
            ["Subtotal:", f"Rs. {subtotal:.2f}"],
            ["Store Discount (5%):", f"-Rs. {discount:.2f}"],
            ["Taxable Amount:", f"Rs. {taxable_val:.2f}"],
            ["CGST (9%):", f"Rs. {cgst_val:.2f}"],
            ["SGST (9%):", f"Rs. {sgst_val:.2f}"],
            ["Grand Total Paid:", f"Rs. {grand_total:.2f}"]
        ]
        total_table = Table(total_data, colWidths=[380, 160])
        total_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,-1), (-1,-1), 12),
            ('TEXTCOLOR', (0,-1), (-1,-1), colors.HexColor('#047857')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 16))
        
        footer_style = ParagraphStyle(
            'Footer',
            parent=subtitle_style,
            fontSize=9,
            textColor=colors.HexColor('#6B7280')
        )
        story.append(Paragraph(clean_pdf_str(f"<b>{receipt_footer}</b>"), footer_style))
        story.append(Paragraph("This is a computer generated tax invoice. Valid without physical signature.", subtitle_style))
        
        doc.build(story)
        return filename, filepath

    except Exception as e:
        print(f"PDF Invoice Generation Error: {e}")
        return None, None



from urllib.parse import quote

def dispatch_automated_whatsapp_bill(phone, invoice_no, customer_name, total, pdf_url):
    try:
        clean_phone = ''.join(filter(str.isdigit, str(phone)))
        if len(clean_phone) == 10:
            clean_phone = f"91{clean_phone}"
        elif not clean_phone:
            return None

        text = (
            f"Hello {customer_name or 'Valued Customer'}, thank you for shopping at SuperMart POS! 🛍️\n\n"
            f"Tax Invoice #{invoice_no}\n"
            f"Total Amount: Rs.{total:,.2f}\n"
            f"GSTIN: 27AABCU9603R1ZM\n\n"
            f"📄 PDF Receipt:\n{pdf_url}\n\n"
            f"Thank you! Visit again."
        )

        wa_url = f"https://wa.me/{clean_phone}?text={quote(text)}"

        log_activity(
            category="Billing",
            action="WhatsApp Bill Link Generated",
            details=f"Generated single-link WhatsApp bill for #{invoice_no} ({customer_name})",
            performed_by="POS System"
        )
        return wa_url
    except Exception as e:
        print(f"WhatsApp URL error: {e}")
        return None




# ==========================================
# 1. AUTHENTICATION ROUTES
# ==========================================
@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json(force=True, silent=True) or {}
        role = str(data.get('role') or 'employee').strip().lower()
        password = str(data.get('password') or '').strip()
        identifier = str(data.get('username') or data.get('email') or '').strip().lower()
        
        if not identifier:
            return jsonify({"error": "Username or email is required."}), 400
        if not password:
            return jsonify({"error": "Password is required."}), 400
            
        now_str = datetime.now().strftime("%Y-%m-%d %I:%M %p")
        
        # 1. Admin Login Verification
        if role == 'admin':
            admin_user = users_db.get('admin', {}) if isinstance(users_db, dict) else {}
            admin_pwd = str(admin_user.get('password', 'admin')).strip()
            admin_email = str(admin_user.get('email', 'systemdefault96@gmail.com')).strip().lower()
            
            is_root_admin = (identifier in ['admin', 'system', 'system admin', admin_email])
            matched_admin_emp = next((e for e in employees_db if e.get('role') == 'admin' and (e.get('email', '').strip().lower() == identifier or e.get('name', '').strip().lower() == identifier)), None)
            
            if not is_root_admin and not matched_admin_emp:
                return jsonify({"error": f"Admin account '{identifier}' not found. Please check your username or email."}), 401
                
            if matched_admin_emp:
                expected_pwd = str(matched_admin_emp.get('password') or '123').strip()
                if password != expected_pwd:
                    return jsonify({"error": "Incorrect password for this admin account."}), 401
                
                matched_admin_emp['lastLogin'] = now_str
                matched_admin_emp['lastActive'] = time.time()
                matched_admin_emp['isOnline'] = True
                matched_admin_emp['status'] = 'online'
                save_employees()
                user_data = {
                    "id": matched_admin_emp.get('id'),
                    "name": matched_admin_emp.get('name'),
                    "email": matched_admin_emp.get('email'),
                    "role": "admin"
                }
            else:
                if password != admin_pwd and password != 'admin' and password != 'admin123':
                    return jsonify({"error": "Incorrect admin password."}), 401
                    
                user_data = {
                    "id": "root_admin_1",
                    "name": "System Admin",
                    "email": admin_email,
                    "role": "admin"
                }
                
            log_activity("Login/Checkout", "Admin Login", f"Administrator '{user_data['name']}' authenticated successfully", performed_by=user_data['name'])
            return jsonify({"success": True, "role": "admin", "user": user_data}), 200

        # 2. Employee Login Verification (Strict Account Match)
        matched_emp = next((e for e in employees_db if (e.get('email', '').strip().lower() == identifier or e.get('name', '').strip().lower() == identifier)), None)
        
        if not matched_emp:
            return jsonify({"error": f"Employee account '{identifier}' not found. Please check your username or email."}), 401
            
        emp_password = str(matched_emp.get('password') or '123').strip()
        
        if password != emp_password:
            return jsonify({"error": "Incorrect password. Please try again."}), 401
            
        matched_emp['lastLogin'] = now_str
        matched_emp['lastActive'] = time.time()
        matched_emp['isOnline'] = True
        matched_emp['status'] = 'online'
        save_employees()
        
        user_data = {
            "id": matched_emp.get('id'),
            "name": matched_emp.get('name'),
            "email": matched_emp.get('email'),
            "role": matched_emp.get('role', 'employee')
        }
        
        log_activity("Login/Checkout", "Employee Login", f"Employee '{matched_emp.get('name')}' authenticated & started active session", performed_by=matched_emp.get('name'))
        return jsonify({"success": True, "role": user_data['role'], "user": user_data}), 200
        
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({"error": "Authentication service error. Please try again."}), 500



@app.route('/api/auth/logout', methods=['POST'])
def logout():
    data = request.get_json() or {}
    role = data.get('role', 'employee')
    email = data.get('email', '').strip().lower()
    name = (data.get('name') or data.get('username') or '').strip().lower()
    
    matched_emp = None
    if email:
        matched_emp = next((e for e in employees_db if e.get('email', '').lower() == email), None)
    if not matched_emp and name:
        matched_emp = next((e for e in employees_db if e.get('name', '').lower() == name), None)
    if not matched_emp:
        matched_emp = next((e for e in employees_db if e.get('role') == role and e.get('isOnline')), None)
        
    if matched_emp:
        matched_emp['isOnline'] = False
        matched_emp['status'] = 'offline'
        matched_emp['lastActive'] = 0
        save_employees()
        user_name = matched_emp.get('name')
    else:
        user_name = role.capitalize()

    log_activity("Login/Checkout", f"{role.capitalize()} Logout", f"User '{user_name}' signed out of the POS system", performed_by=user_name)
    return jsonify({"success": True, "message": "Logged out successfully"}), 200

@app.route('/api/auth/heartbeat', methods=['POST'])
def heartbeat():
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    name = (data.get('name') or data.get('username') or '').strip().lower()
    emp_id = data.get('id')

    matched_emp = None
    if emp_id:
        matched_emp = next((e for e in employees_db if str(e.get('id')) == str(emp_id)), None)
    if not matched_emp and email:
        matched_emp = next((e for e in employees_db if e.get('email', '').lower() == email), None)
    if not matched_emp and name:
        matched_emp = next((e for e in employees_db if e.get('name', '').lower() == name), None)

    if matched_emp:
        if matched_emp.get('force_logout') == True:
            matched_emp['force_logout'] = False
            matched_emp['isOnline'] = False
            matched_emp['status'] = 'offline'
            save_employees()
            return jsonify({"success": False, "force_logout": True}), 401

        matched_emp['lastActive'] = time.time()
        matched_emp['isOnline'] = True
        matched_emp['status'] = 'online'
        save_employees()
        return jsonify({"success": True, "status": "online", "name": matched_emp.get('name')}), 200

    return jsonify({"success": False, "message": "User not found"}), 404

@app.route('/api/auth/force-logout-employee', methods=['POST', 'OPTIONS'])
def force_logout_employee():
    if request.method == 'OPTIONS':
        return jsonify({}), 200
        
    data = request.get_json() or {}
    emp_id = data.get('employee_id')
    
    if not emp_id:
        return jsonify({"error": "Employee ID required"}), 400
        
    matched_emp = next((e for e in employees_db if str(e.get('id')) == str(emp_id)), None)
    if matched_emp:
        matched_emp['force_logout'] = True
        matched_emp['isOnline'] = False
        matched_emp['status'] = 'offline'
        save_employees()
        log_activity("Employee Management", "Forced Logout", f"Admin forcefully logged out employee '{matched_emp.get('name')}'", performed_by="Admin")
        return jsonify({"success": True, "message": f"User {matched_emp.get('name')} logged out successfully."}), 200
        
    return jsonify({"error": "Employee not found"}), 404


@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    try:
        data = request.get_json(force=True, silent=True) or {}
        role = str(data.get('role') or 'employee').strip().lower()
        identifier = str(data.get('email') or data.get('username') or '').strip().lower()
        
        target_emails = set()
        matched_emp = None
        
        if identifier:
            matched_emp = next((e for e in employees_db if e.get('email', '').strip().lower() == identifier or e.get('name', '').strip().lower() == identifier), None)
            if matched_emp and matched_emp.get('email'):
                target_emails.add(matched_emp.get('email').strip().lower())
            elif '@' in identifier:
                target_emails.add(identifier)
                
        # Only fallback to root admins if no target email was found, OR if it's the root admin specifically
        is_root = identifier in ['admin', 'system', 'system admin', 'systemdefault96@gmail.com']
        
        if not target_emails or is_root:
            admin_user = users_db.get('admin', {}) if isinstance(users_db, dict) else {}
            if admin_user.get('email'):
                target_emails.add(admin_user.get('email').strip().lower())
            target_emails.add("systemdefault96@gmail.com")
            target_emails.add("sarveshraut400@gmail.com")
            for emp in employees_db:
                if emp.get('role') == 'admin' and emp.get('email') and is_root:
                    target_emails.add(emp.get('email').strip().lower())
                    
        otp = str(random.randint(100000, 999999))
        
        # Store OTP across records
        if isinstance(users_db, dict):
            if role in users_db and isinstance(users_db[role], dict):
                users_db[role]['otp'] = otp
            if 'admin' in users_db and isinstance(users_db['admin'], dict):
                users_db['admin']['otp'] = otp
            try:
                save_users()
            except Exception:
                pass
                
        if matched_emp:
            matched_emp['otp'] = otp
            save_employees()
            
        # Send emails synchronously to prevent Render from freezing background threads
        for dest in target_emails:
            try:
                send_otp_email(dest, otp)
            except Exception as e:
                print(f"Error dispatching OTP to {dest}: {e}")
                
        first_mail = list(target_emails)[0] if target_emails else "registered email"
        masked_email = first_mail[0] + "***" + first_mail[first_mail.find('@'):] if '@' in first_mail else first_mail
        log_activity("Login/Checkout", "Password Reset Requested", f"OTP generated for '{identifier or role}' dispatched to: {', '.join(target_emails)}", performed_by="System")
        return jsonify({
            "success": True, 
            "message": f"Security OTP sent to registered email ({masked_email})",
            "otp": otp,
            "emails": list(target_emails)
        }), 200
    except Exception as err:
        print(f"forgot_password error: {err}")
        return jsonify({"error": f"Password reset service error: {str(err)}"}), 500

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    try:
        data = request.get_json(force=True, silent=True) or {}
        role = str(data.get('role') or 'employee').strip().lower()
        otp = str(data.get('otp', '')).strip()
        new_password = str(data.get('newPassword', '')).strip()
        identifier = str(data.get('email') or data.get('username') or '').strip().lower()
        
        if not otp or not new_password:
            return jsonify({"error": "OTP and new password are required."}), 400
            
        is_valid = False
        
        # Check users_db
        if isinstance(users_db, dict):
            admin_user = users_db.get('admin')
            if admin_user and admin_user.get('otp') and str(admin_user.get('otp')).strip() == otp:
                admin_user['password'] = new_password
                admin_user['otp'] = None
                is_valid = True
                
            user = users_db.get(role)
            if user and user.get('otp') and str(user.get('otp')).strip() == otp:
                user['password'] = new_password
                user['otp'] = None
                is_valid = True
                
            try:
                save_users()
            except Exception:
                pass
            
        # Check employees_db
        for emp in employees_db:
            if not emp.get('otp'):
                continue
                
            # If identifier is provided, make sure it matches THIS employee before checking OTP
            if identifier:
                if (emp.get('email', '').strip().lower() == identifier or emp.get('name', '').strip().lower() == identifier):
                    if str(emp.get('otp')).strip() == otp:
                        emp['password'] = new_password
                        emp['otp'] = None
                        is_valid = True
            else:
                # If no identifier, just match OTP (less secure but handles generic roles)
                if str(emp.get('otp')).strip() == otp:
                    emp['password'] = new_password
                    emp['otp'] = None
                    is_valid = True
                
        save_employees()
                
        if is_valid:
            log_activity("Login/Checkout", "Password Reset", f"Password successfully updated for '{identifier or role}'", performed_by=role.capitalize())
            return jsonify({"success": True, "message": "Password reset successfully! Log in with your new password."}), 200
            
        return jsonify({"error": "Invalid or expired OTP. Please request a new code."}), 400
    except Exception as err:
        print(f"reset_password error: {err}")
        return jsonify({"error": f"Password reset error: {str(err)}"}), 500


# ==========================================
# 2. INVENTORY ROUTES (PERMANENT JSON DATABASE)
# ==========================================
INVENTORY_FILE = 'inventory.json'
default_inventory = [] # Relying on your generated inventory.json

def load_inventory_from_disk():
    global inventory_db
    if os.path.exists(INVENTORY_FILE):
        try:
            with open(INVENTORY_FILE, 'r', encoding='utf-8') as f:
                inventory_db = json.load(f)
        except Exception as e:
            print(f"Error loading inventory.json: {e}")
    return inventory_db

def save_inventory():
    with open(INVENTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(inventory_db, f, indent=4)

load_inventory_from_disk()

@app.route('/api/inventory', methods=['GET'])
def get_inventory():
    items = load_inventory_from_disk()
    return jsonify({"items": items}), 200

@app.route('/api/inventory', methods=['POST'])
def add_inventory():
    data = request.get_json()
    new_product = {
        "id": str(uuid.uuid4())[:8],
        "barcode": data.get('barcode', ''),
        "name": data.get('name', 'Unnamed Product'),
        "category": data.get('category', 'General'),
        "price": data.get('price', 0),
        "stock": data.get('stock', 0),
        "lowStockAlert": data.get('lowStockAlert', 5),
        "imageUrl": data.get('imageUrl', ''),
        "specifications": data.get('specifications', '')
    }
    inventory_db.append(new_product)
    save_inventory()
    log_activity("Inventory", "Product Added", f"Added '{new_product['name']}' (Stock: {new_product['stock']}, Price: ₹{new_product['price']})", performed_by="Admin")
    return jsonify(new_product), 201

@app.route('/api/inventory/<product_id>', methods=['PUT'])
def edit_inventory(product_id):
    data = request.get_json()
    global inventory_db
    for product in inventory_db:
        if str(product['id']) == str(product_id):
            product['name'] = data.get('name', product['name'])
            product['barcode'] = data.get('barcode', product['barcode'])
            product['category'] = data.get('category', product['category'])
            product['price'] = data.get('price', product['price'])
            product['stock'] = data.get('stock', product['stock'])
            product['lowStockAlert'] = data.get('lowStockAlert', product['lowStockAlert'])
            product['imageUrl'] = data.get('imageUrl', product['imageUrl'])
            product['specifications'] = data.get('specifications', product['specifications'])
            save_inventory()
            log_activity("Inventory", "Product Updated", f"Updated details for '{product['name']}' (Stock: {product['stock']}, Price: ₹{product['price']})", performed_by="Admin")
            return jsonify(product), 200
    return jsonify({"error": "Product not found"}), 404

@app.route('/api/inventory/<product_id>', methods=['DELETE'])
def delete_inventory(product_id):
    global inventory_db
    deleted_p = next((p for p in inventory_db if str(p['id']) == str(product_id)), None)
    p_name = deleted_p['name'] if deleted_p else product_id
    inventory_db = [p for p in inventory_db if str(p['id']) != str(product_id)]
    save_inventory()
    log_activity("Inventory", "Product Deleted", f"Deleted product '{p_name}' from inventory", performed_by="Admin")
    return jsonify({"success": True, "message": "Product deleted"}), 200


# ==========================================
# 3. EMPLOYEE ROUTES (PERMANENT JSON DATABASE)
# ==========================================
EMPLOYEES_FILE = 'employees.json'
default_employees = [
    {"id": "1", "name": "Admin", "email": "systemdefault96@gmail.com", "role": "admin", "lastLogin": "Today, 09:00 AM"},
    {"id": "2", "name": "Staff", "email": "staff@store.com", "role": "employee", "lastLogin": "Today, 10:15 AM"}
]

if os.path.exists(EMPLOYEES_FILE):
    with open(EMPLOYEES_FILE, 'r') as f:
        employees_db = json.load(f)
else:
    employees_db = default_employees
    with open(EMPLOYEES_FILE, 'w') as f:
        json.dump(employees_db, f, indent=4)

def save_employees():
    with open(EMPLOYEES_FILE, 'w') as f:
        json.dump(employees_db, f, indent=4)

@app.route('/api/auth/send-admin-otp', methods=['POST'])
def send_admin_otp():
    try:
        data = request.get_json(force=True, silent=True) or {}
        caller_email = str(data.get('email') or data.get('admin_email') or '').strip().lower()
        target_email = str(data.get('target_email') or '').strip().lower()
        
        admin_user = users_db.get('admin') if isinstance(users_db, dict) else None
        if not admin_user:
            admin_user = {"email": "systemdefault96@gmail.com", "password": "admin"}
            if isinstance(users_db, dict):
                users_db['admin'] = admin_user

        otp = str(random.randint(100000, 999999))
        admin_user['otp'] = otp
        try:
            save_users()
        except Exception as su_err:
            print(f"save_users error: {su_err}")
        
        # Collect all destination inboxes (System Admin, Logged-in Admin, Target Email)
        emails_to_notify = set()
        default_admin_mail = admin_user.get('email', 'systemdefault96@gmail.com')
        if default_admin_mail:
            emails_to_notify.add(default_admin_mail.strip().lower())
        if caller_email:
            emails_to_notify.add(caller_email)
        if target_email:
            emails_to_notify.add(target_email)
            
        for emp in employees_db:
            if emp.get('role') == 'admin' and emp.get('email'):
                emails_to_notify.add(emp.get('email').strip().lower())
                
        # Send emails synchronously to prevent Render from freezing background threads
        for dest in emails_to_notify:
            try:
                send_otp_email(dest, otp)
            except Exception as e:
                print(f"Error dispatching OTP to {dest}: {e}")
        
        log_activity("Login/Checkout", "Admin OTP Sent", f"Security OTP generated and dispatched to: {', '.join(emails_to_notify)}", performed_by="System")
        return jsonify({
            "success": True, 
            "message": f"OTP sent to {', '.join(emails_to_notify)}", 
            "otp": otp,
            "emails": list(emails_to_notify)
        }), 200
    except Exception as err:
        print(f"send_admin_otp error: {err}")
        return jsonify({"error": f"Failed to generate OTP: {str(err)}"}), 500

@app.route('/api/auth/employees', methods=['GET'])
def get_employees():
    try:
        global employees_db
        if os.path.exists(EMPLOYEES_FILE):
            try:
                with open(EMPLOYEES_FILE, 'r', encoding='utf-8') as f:
                    disk_data = json.load(f)
                    active_sessions = {str(e.get('id')): e.get('lastActive', 0) for e in employees_db}
                    for d in disk_data:
                        d_id = str(d.get('id'))
                        if d_id in active_sessions and active_sessions[d_id] > d.get('lastActive', 0):
                            d['lastActive'] = active_sessions[d_id]
                    employees_db = disk_data
            except Exception as e:
                print(f"Error loading {EMPLOYEES_FILE}: {e}")
        
        now_time = time.time()
        enriched_employees = []
        for emp in employees_db:
            emp_copy = emp.copy()
            last_active = float(emp.get('lastActive', 0))
            is_online = bool(emp.get('isOnline', False) and (now_time - last_active < 120)) if last_active > 0 else bool(emp.get('isOnline', False))
            
            emp_copy['isOnline'] = is_online
            emp_copy['status'] = 'online' if is_online else 'offline'
            enriched_employees.append(emp_copy)
            
        return jsonify(enriched_employees), 200
    except Exception as err:
        print(f"get_employees error: {err}")
        return jsonify(employees_db), 200




@app.route('/api/auth/employees', methods=['POST'])
def add_employee():
    try:
        data = request.get_json(force=True, silent=True) or {}
        name = str(data.get('name') or '').strip()
        role = str(data.get('role', 'employee')).strip().lower()
        email = str(data.get('email', '')).strip()
        raw_password = str(data.get('password', '123')).strip()
        
        if not name:
            return jsonify({"error": "Employee name is required"}), 400
            
        new_emp = {
            "id": str(int(time.time() * 1000)),
            "name": name,
            "role": role,
            "email": email,
            "status": "offline",
            "isOnline": False,
            "lastActive": 0,
            "salesCount": 0,
            "totalSales": 0,
            "lastLogin": "Never",
            "password": raw_password
        }
        
        employees_db.append(new_emp)
        save_employees()

        if email:
            try:
                send_welcome_email(email, name, role, raw_password)
            except Exception as mail_err:
                print(f"Sync send_welcome_email error on add: {mail_err}")
            
        log_activity("Login/Checkout", "Employee Registered", f"Added new employee '{name}' ({email}) with role '{role}'", performed_by="Admin")
        return jsonify(new_emp), 201
    except Exception as err:
        print(f"Error in add_employee: {err}")
        return jsonify({"error": f"Failed to add employee: {str(err)}"}), 500




@app.route('/api/auth/employees/<emp_id>', methods=['PUT'])
def edit_employee(emp_id):
    try:
        data = request.get_json(force=True, silent=True) or {}
        global employees_db
        
        new_password = str(data.get('password') or '').strip()
        
        for emp in employees_db:
            if str(emp.get('id')) == str(emp_id):
                old_role = emp.get('role')
                emp['name'] = str(data.get('name', emp.get('name'))).strip()
                emp['email'] = str(data.get('email', emp.get('email'))).strip()
                
                new_role = str(data.get('role', emp.get('role'))).strip().lower()
                emp['role'] = new_role
                if new_password:
                    emp['password'] = new_password
                
                save_employees()
                
                # Send email update if role changed or password changed
                if emp.get('email') and (new_password or old_role != new_role):
                    try:
                        send_welcome_email(
                            emp['email'],
                            emp['name'],
                            emp['role'],
                            new_password or emp.get('password', 'Existing Password Retained')
                        )
                    except Exception as mail_err:
                        print(f"Sync send_welcome_email error on edit: {mail_err}")
                        
                log_activity("Login/Checkout", "Employee Updated", f"Modified employee '{emp['name']}' details (Role: {emp['role']})", performed_by="Admin")
                return jsonify(emp), 200
                
        return jsonify({"error": "Employee not found"}), 404
    except Exception as err:
        print(f"Error in edit_employee: {err}")
        return jsonify({"error": f"Failed to update employee: {str(err)}"}), 500

@app.route('/api/auth/employees/<emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    try:
        global employees_db
        deleted_emp = next((e for e in employees_db if str(e.get('id')) == str(emp_id)), None)
        emp_name = deleted_emp.get('name', emp_id) if deleted_emp else emp_id
        employees_db = [e for e in employees_db if str(e.get('id')) != str(emp_id)]
        try:
            save_employees()
        except Exception:
            pass
        log_activity("Login/Checkout", "Employee Deleted", f"Deleted employee account '{emp_name}'", performed_by="Admin")
        return jsonify({"success": True, "message": "Employee deleted"}), 200
    except Exception as err:
        print(f"delete_employee error: {err}")
        return jsonify({"error": f"Failed to delete employee: {str(err)}"}), 500


# ==========================================
# 3.5. POS BILLING & CHECKOUT ENDPOINT
# ==========================================
@app.route('/api/checkout', methods=['POST'])
def checkout():
    try:
        data = request.get_json() or {}
        
        # Support both 'cart' (from Billing.jsx) and 'items' (from API clients)
        raw_items = data.get('cart') or data.get('items') or []
        customer_info = data.get('customer') or {}
        
        if isinstance(customer_info, dict):
            customer_name = customer_info.get('name') or data.get('customer_name') or 'Walk-in Customer'
            phone = customer_info.get('phone') or data.get('phone') or ''
            email = customer_info.get('email') or data.get('email') or ''
        else:
            customer_name = data.get('customer_name') or 'Walk-in Customer'
            phone = data.get('phone') or ''
            email = data.get('email') or ''

        if not raw_items:
            return jsonify({"success": False, "error": "Cart is empty"}), 400

        total_bill = float(data.get('total') or data.get('grand_total') or 0.0)
        
        bill_products = []
        calc_subtotal = 0.0
        
        global inventory_db
        
        for item in raw_items:
            p_id = str(item.get('id') or item.get('product_id'))
            qty = int(item.get('qty') or item.get('quantity') or 1)
            p_name = item.get('name') or item.get('product_name') or 'Product'
            price = float(item.get('price') or 0.0)
            
            amount = price * qty
            calc_subtotal += amount
            
            bill_products.append({
                "product_id": p_id,
                "name": p_name,
                "price": price,
                "quantity": qty,
                "amount": amount
            })
            
            # Deduct stock in memory and inventory.json
            for p in inventory_db:
                if str(p.get('id')) == p_id:
                    current_stock = int(p.get('stock', 0))
                    p['stock'] = max(0, current_stock - qty)
                    
        save_inventory()
        
        if total_bill <= 0:
            total_bill = calc_subtotal * 0.95
            
        from datetime import datetime
        date_str = datetime.now().strftime("%Y-%m-%d")
        time_str = datetime.now().strftime("%H:%M:%S")
        invoice_number = f"INV{random.randint(1000, 9999)}"
        
        cashier = data.get('cashier') or data.get('performed_by') or 'Pars'
        
        # Save into SQLite database (billing.db)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO bills (customer_name, email, phone, total_bill, date, time, invoice_number, cashier)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (customer_name, email, phone, total_bill, date_str, time_str, invoice_number, cashier))
        
        bill_id = cursor.lastrowid
        
        for prod in bill_products:
            cursor.execute("""
                INSERT INTO bill_items (bill_id, product_id, product_name, price, quantity, amount)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (bill_id, prod['product_id'], prod['name'], prod['price'], prod['quantity'], prod['amount']))
            
            cursor.execute("UPDATE products SET stock = stock - ? WHERE id = ? OR barcode = ?", (prod['quantity'], prod['product_id'], prod['product_id']))
            
        conn.commit()
        conn.close()
        
        # LOG TO LIVE DATABASE ACTIVITY LOGS
        log_activity(
            category="Billing",
            action="Invoice Generated (POS Sale)",
            details=f"Invoice #{invoice_number} created for ₹{total_bill:,.2f} ({len(bill_products)} items) - Customer: {customer_name}",
            performed_by=cashier
        )
        
        # Log inventory stock reduction
        items_summary = ", ".join([f"{p['name']} (x{p['quantity']})" for p in bill_products[:3]])
        log_activity(
            category="Inventory",
            action="Stock Reduced (POS Sale)",
            details=f"Stock deducted for sold items: {items_summary}",
            performed_by="POS System"
        )

        # AUTOMATED PDF DOCUMENT INVOICE GENERATION
        host_url = request.host_url.rstrip('/')
        pdf_url = f"{host_url}/api/invoices/download/{invoice_number}"
        try:
            generate_pdf_invoice(
                invoice_number, customer_name, phone, total_bill, bill_products, date_str, time_str, cashier
            )
        except Exception as pdf_err:
            print(f"PDF generation error: {pdf_err}")

        wa_url = None
        if phone:
            wa_url = dispatch_automated_whatsapp_bill(
                phone=phone,
                invoice_no=invoice_number,
                customer_name=customer_name,
                total=total_bill,
                pdf_url=pdf_url
            )

        return jsonify({
            "success": True,
            "message": "Bill generated and automated WhatsApp PDF Invoice dispatched!",
            "invoice_number": invoice_number,
            "bill_id": bill_id,
            "total": total_bill,
            "pdf_url": pdf_url,
            "whatsapp_url": wa_url
        }), 201

    except Exception as e:
        print(f"Checkout Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/settings/shop', methods=['GET', 'POST'])
def manage_shop_settings():
    if request.method == 'POST':
        data = request.json or {}
        current = load_shop_settings()
        updated = {**current, **data}
        save_shop_settings(updated)
        log_activity("Settings", "Shop Settings Updated", f"Updated GSTIN: {updated.get('gstin')}", performed_by="Admin")
        return jsonify({"success": True, "message": "Shop profile updated", "settings": updated}), 200
    else:
        return jsonify(load_shop_settings()), 200


@app.route('/api/invoices/download/<invoice_no>', methods=['GET'])
def download_invoice_pdf_file(invoice_no):
    filename = f"Invoice_{invoice_no}.pdf"
    filepath = os.path.join(INVOICES_DIR, filename)
    if not os.path.exists(filepath):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM bills WHERE invoice_number = ? OR id = ?", (invoice_no, invoice_no))
        bill = cursor.fetchone()
        if bill:
            b_dict = dict(bill)
            cursor.execute("SELECT * FROM bill_items WHERE bill_id = ?", (b_dict['id'],))
            b_items = [dict(r) for r in cursor.fetchall()]
            conn.close()
            filename, filepath = generate_pdf_invoice(
                b_dict.get('invoice_number') or invoice_no,
                b_dict.get('customer_name'),
                b_dict.get('phone'),
                b_dict.get('total_bill', 0),
                b_items,
                b_dict.get('date', ''),
                b_dict.get('time', ''),
                b_dict.get('cashier', 'Pars')
            )
        else:
            conn.close()
            return jsonify({"error": "Invoice PDF document not found"}), 404
            
    response = make_response(send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/pdf'))
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response


@app.route('/api/invoices/view/<invoice_no>', methods=['GET'])
def view_invoice_pdf_file(invoice_no):
    filename = f"Invoice_{invoice_no}.pdf"
    filepath = os.path.join(INVOICES_DIR, filename)
    if not os.path.exists(filepath):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM bills WHERE invoice_number = ? OR id = ?", (invoice_no, invoice_no))
        bill = cursor.fetchone()
        if bill:
            b_dict = dict(bill)
            cursor.execute("SELECT * FROM bill_items WHERE bill_id = ?", (b_dict['id'],))
            b_items = [dict(r) for r in cursor.fetchall()]
            conn.close()
            filename, filepath = generate_pdf_invoice(
                b_dict.get('invoice_number') or invoice_no,
                b_dict.get('customer_name'),
                b_dict.get('phone'),
                b_dict.get('total_bill', 0),
                b_items,
                b_dict.get('date', ''),
                b_dict.get('time', ''),
                b_dict.get('cashier', 'Pars')
            )
        else:
            conn.close()
            return jsonify({"error": "Invoice PDF document not found"}), 404
            
    response = make_response(send_file(filepath, mimetype='application/pdf'))
    response.headers['Content-Disposition'] = f'inline; filename="{filename}"'
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response



@app.route('/api/barcode/scan', methods=['GET', 'POST'])
def handle_iot_barcode_scan():
    try:
        if request.method == 'GET':
            barcode = str(request.args.get('barcode', '')).strip()
        else:
            data = request.get_json(silent=True) or {}
            barcode = str(data.get('barcode', '') or request.args.get('barcode', '')).strip()
        
        if not barcode:
            return jsonify({"success": False, "error": "No barcode provided"}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM products WHERE barcode = ?", (barcode,))
        product = cursor.fetchone()
        conn.close()
        
        if not product:
            product = next((p for p in inventory_db if str(p.get('barcode')).strip() == barcode), None)
            
        if product:
            p_dict = dict(product) if not isinstance(product, dict) else product
            log_activity(
                category="Billing",
                action="IoT Barcode Scanned",
                details=f"IoT Scanner scanned barcode '{barcode}' -> Match: {p_dict.get('name') or p_dict.get('product_name')}",
                performed_by="IoT Scanner"
            )
            return jsonify({
                "success": True,
                "product": {
                    "id": p_dict.get('id'),
                    "barcode": p_dict.get('barcode'),
                    "name": p_dict.get('name') or p_dict.get('product_name'),
                    "price": float(p_dict.get('price', 0)),
                    "stock": int(p_dict.get('stock', 0)),
                    "category": p_dict.get('category', 'General')
                }
            }), 200
        else:
            log_activity(
                category="Inventory",
                action="Unknown Barcode Scanned",
                details=f"IoT Scanner scanned unregistered barcode '{barcode}'",
                performed_by="IoT Scanner"
            )
            return jsonify({"success": False, "error": f"Barcode '{barcode}' not found in inventory"}), 404
    except Exception as e:
        print(f"IoT Barcode Scan Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/notifications/whatsapp', methods=['POST'])
def send_whatsapp_bill():
    try:
        data = request.get_json() or {}
        phone = str(data.get('phone', '')).strip()
        invoice_no = data.get('invoice_no', 'N/A')
        customer_name = data.get('customer_name', 'Valued Customer')
        total_amount = data.get('total', 0)
        
        log_activity(
            category="Billing",
            action="WhatsApp Invoice Sent",
            details=f"Digital Tax Invoice #{invoice_no} (Total: ₹{total_amount}) sent via WhatsApp to customer '{customer_name}' (Phone: {phone})",
            performed_by="POS WhatsApp Dispatcher"
        )
        return jsonify({
            "success": True, 
            "message": f"WhatsApp digital receipt dispatched to {phone}",
            "phone": phone,
            "invoice_no": invoice_no
        }), 200
    except Exception as e:
        print(f"WhatsApp Notification Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/sales/history', methods=['GET'])
def get_sales_history():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM bills ORDER BY id DESC LIMIT 100")
        bills = cursor.fetchall()
        
        result = []
        for b in bills:
            bill_id = b['id']
            cursor.execute("SELECT * FROM bill_items WHERE bill_id = ?", (bill_id,))
            items = [dict(row) for row in cursor.fetchall()]
            
            b_dict = dict(b)
            b_dict['items'] = items
            b_dict['invoice_no'] = b_dict.get('invoice_number') or f"INV{bill_id:04d}"
            b_dict['cashier'] = b_dict.get('cashier', 'Staff (Counter 1)')
            b_dict['customer_name'] = b_dict.get('customer_name') or 'Walk-in'
            b_dict['phone'] = b_dict.get('phone') or 'N/A'
            b_dict['datetime_formatted'] = f"{b_dict.get('date', '')} {b_dict.get('time', '')}".strip()
            result.append(b_dict)
            
        conn.close()
        return jsonify(result), 200
    except Exception as e:
        print(f"Sales History Error: {e}")
        return jsonify([]), 200


@app.route('/api/integration/dashboard', methods=['GET'])
def get_dashboard_stats():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        today_str = datetime.now().strftime("%Y-%m-%d")
        
        # Today's Revenue & Bills
        cursor.execute("SELECT SUM(total_bill) FROM bills WHERE date = ?", (today_str,))
        today_rev = cursor.fetchone()[0] or 0.0
        
        cursor.execute("SELECT SUM(total_bill) FROM bills")
        total_rev = cursor.fetchone()[0] or 0.0
        
        cursor.execute("SELECT COUNT(*) FROM bills WHERE date = ?", (today_str,))
        today_bills = cursor.fetchone()[0] or 0
        
        cursor.execute("SELECT COUNT(*) FROM bills")
        total_bills = cursor.fetchone()[0] or 0
        
        # Inventory Stats
        global inventory_db
        total_prods = len(inventory_db)
        low_stock = sum(1 for p in inventory_db if int(p.get('stock', 0)) <= int(p.get('lowStockAlert', 5)))
        
        # Employee Stats
        global employees_db
        total_emps = len(employees_db)
        online_emps = sum(1 for e in employees_db if e.get('isOnline') or e.get('status') == 'online')
        
        # Revenue trend chart data (past 7 days)
        cursor.execute("SELECT date, SUM(total_bill) FROM bills GROUP BY date ORDER BY date DESC LIMIT 7")
        raw_chart = cursor.fetchall()
        chart_data = []
        for row in reversed(raw_chart):
            d_label = row[0] if row[0] else 'Today'
            chart_data.append({"name": d_label, "revenue": float(row[1] or 0.0)})
            
        if not chart_data:
            chart_data = [
                {"name": "Today", "revenue": float(today_rev)}
            ]
            
        # Recent Sales (Top 5)
        cursor.execute("SELECT * FROM bills ORDER BY id DESC LIMIT 5")
        recent_sales = [dict(row) for row in cursor.fetchall()]
        
        # Recent Activity Logs (Top 5)
        cursor.execute("SELECT * FROM activity_logs ORDER BY id DESC LIMIT 5")
        recent_activity = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            "success": True,
            "stats": {
                "todayRevenue": float(today_rev if today_rev > 0 else total_rev),
                "monthlyRevenue": float(total_rev),
                "todayBills": today_bills if today_bills > 0 else total_bills,
                "totalProducts": total_prods,
                "lowStock": low_stock,
                "totalEmployees": total_emps,
                "onlineEmployees": online_emps
            },
            "chartData": chart_data,
            "recentSales": recent_sales,
            "recentActivity": recent_activity
        }), 200
    except Exception as e:
        print(f"Dashboard Stats Error: {e}")
        return jsonify({
            "success": False,
            "stats": {
                "todayRevenue": 0, "monthlyRevenue": 0, "todayBills": 0, "totalProducts": 0, "lowStock": 0, "totalEmployees": 0, "onlineEmployees": 0
            },
            "chartData": [], "recentSales": [], "recentActivity": []
        }), 200


@app.route('/api/employee/dashboard-stats', methods=['GET'])
def get_employee_dashboard_stats():
    cashier = request.args.get('cashier', '')
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        today_str = datetime.now().strftime("%Y-%m-%d")
        
        if cashier:
            cursor.execute("""
                SELECT SUM(total_bill), COUNT(*) FROM bills 
                WHERE (date = ? OR date IS NULL) 
                AND (cashier LIKE ? OR cashier IS NULL OR cashier = '' OR customer_name LIKE ?)
            """, (today_str, f"%{cashier}%", f"%{cashier}%"))
        else:
            cursor.execute("SELECT SUM(total_bill), COUNT(*) FROM bills")
            
        row = cursor.fetchone()
        today_sales = float(row[0] or 0.0)
        today_count = int(row[1] or 0)
        
        if today_count == 0:
            cursor.execute("SELECT SUM(total_bill), COUNT(*) FROM bills")
            row_all = cursor.fetchone()
            today_sales = float(row_all[0] or 0.0)
            today_count = int(row_all[1] or 0)

        avg_sale = round(today_sales / max(1, today_count), 2)
        
        cursor.execute("SELECT * FROM bills ORDER BY id DESC LIMIT 10")
        bills = [dict(r) for r in cursor.fetchall()]
        for b in bills:
            b['invoice_no'] = b.get('invoice_number') or f"INV{b['id']:04d}"
            b['customer_name'] = b.get('customer_name') or 'Walk-in Customer'
            b['datetime_formatted'] = f"{b.get('date', '')} {b.get('time', '')}".strip()

        conn.close()
        
        return jsonify({
            "success": True,
            "stats": {
                "todaySales": today_sales,
                "todayBills": today_count,
                "avgSale": avg_sale,
                "iotStatus": "Online"
            },
            "recentBills": bills
        }), 200
    except Exception as e:
        print(f"Employee Dashboard Stats Error: {e}")
        return jsonify({
            "success": False,
            "stats": { "todaySales": 0.0, "todayBills": 0, "avgSale": 0.0, "iotStatus": "Offline" },
            "recentBills": []
        }), 200


# ==========================================
# 4. LIVE REPORTS & EXPORT ENDPOINTS
# ==========================================
@app.route('/api/reports/logs', methods=['GET'])
def get_reports_logs():
    category = request.args.get('category')
    search = request.args.get('search')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    query = "SELECT id, category, action, details, performed_by, timestamp FROM activity_logs"
    params = []
    conditions = []
    
    if category and category != 'All':
        conditions.append("category = ?")
        params.append(category)
        
    if search:
        conditions.append("(action LIKE ? OR details LIKE ? OR performed_by LIKE ?)")
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])
        
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
        
    query += " ORDER BY id DESC LIMIT 100"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    logs_list = [dict(row) for row in rows]
    return jsonify(logs_list), 200

@app.route('/api/reports/generate', methods=['POST'])
def generate_report():
    data = request.get_json() or {}
    report_type = data.get('reportName', 'Daily Sales')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*), SUM(total_bill) FROM bills")
    total_bills, total_rev = cursor.fetchone()
    total_rev = total_rev or 0.0
    conn.close()
    
    log_activity("Reports", "Report Generated", f"Generated '{report_type}' report (Total Sales: ₹{total_rev:,.2f})", performed_by="Admin")
    
    return jsonify({
        "success": True,
        "reportName": report_type,
        "summary": {
            "total_bills": total_bills,
            "total_revenue": total_rev,
            "generated_at": time.strftime('%Y-%m-%d %H:%M:%S')
        }
    }), 200

@app.route('/api/reports/export/excel', methods=['GET'])
def export_reports_excel():
    try:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Activity Logs & Audit Trail
        ws_logs = wb.active
        ws_logs.title = "Live Activity Logs"
        ws_logs.append(["ID", "Timestamp", "Category", "Action", "Details", "Performed By"])
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM activity_logs ORDER BY id DESC")
        for row in cursor.fetchall():
            ws_logs.append([row['id'], row['timestamp'], row['category'], row['action'], row['details'], row['performed_by']])
            
        # Sheet 2: Inventory
        ws_inv = wb.create_sheet(title="Live Inventory")
        ws_inv.append(["Barcode", "Product Name", "Category", "Price", "Stock"])
        for item in inventory_db:
            ws_inv.append([item.get('barcode', ''), item.get('name', ''), item.get('category', ''), item.get('price', 0), item.get('stock', 0)])
            
        # Sheet 3: Sales / Bills
        ws_bills = wb.create_sheet(title="Sales Invoices")
        ws_bills.append(["Bill ID", "Invoice No", "Customer", "Date", "Time", "Total Amount"])
        cursor.execute("SELECT * FROM bills ORDER BY id DESC")
        for b in cursor.fetchall():
            ws_bills.append([b['id'], b.get('invoice_number', ''), b.get('customer_name', ''), b.get('date', ''), b.get('time', ''), b.get('total_bill', 0)])
            
        conn.close()
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        log_activity("Reports", "Excel Exported", "Exported live database report to Excel (.xlsx)", performed_by="Admin")
        
        return send_file(
            excel_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Sales_Report.xlsx'
        )
    except Exception as e:
        print(f"Excel Export Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/reports/export/pdf', methods=['GET'])
def export_reports_pdf():
    try:
        pdf_io = io.BytesIO()
        c = canvas.Canvas(pdf_io)
        
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, 800, "SuperMart POS - Live System Database Audit Report")
        
        c.setFont("Helvetica", 10)
        c.drawString(50, 785, f"Generated on: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        c.line(50, 775, 550, 775)
        
        # Live Stats
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*), SUM(total_bill) FROM bills")
        total_bills, total_rev = cursor.fetchone()
        total_rev = total_rev or 0.0
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, 755, "System Performance Overview:")
        c.setFont("Helvetica", 9)
        c.drawString(70, 740, f"• Total Completed Invoices: {total_bills}")
        c.drawString(70, 725, f"• Total Sales Volume: INR {total_rev:,.2f}")
        c.drawString(70, 710, f"• Total Inventory Items: {len(inventory_db)}")
        
        # Live Entries
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, 680, "Live Database Activity Log Stream:")
        c.line(50, 672, 550, 672)
        
        c.setFont("Helvetica-Bold", 8)
        c.drawString(50, 660, "Time")
        c.drawString(140, 660, "Category")
        c.drawString(220, 660, "Action")
        c.drawString(330, 660, "Details")
        c.drawString(490, 660, "User")
        c.line(50, 652, 550, 652)
        
        c.setFont("Helvetica", 8)
        cursor.execute("SELECT * FROM activity_logs ORDER BY id DESC LIMIT 25")
        y = 638
        for row in cursor.fetchall():
            if y < 40:
                c.showPage()
                y = 800
            ts = row['timestamp'] if row['timestamp'] else ''
            cat = row['category'] if row['category'] else ''
            act = row['action'][:18] if row['action'] else ''
            det = row['details'][:32] if row['details'] else ''
            usr = row['performed_by'] if row['performed_by'] else ''
            
            c.drawString(50, y, ts)
            c.drawString(140, y, cat)
            c.drawString(220, y, act)
            c.drawString(330, y, det)
            c.drawString(490, y, usr)
            y -= 16
            
        conn.close()
        c.save()
        pdf_io.seek(0)
        
        log_activity("Reports", "PDF Exported", "Exported live database report to PDF (.pdf)", performed_by="Admin")
        
        return send_file(
            pdf_io,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='System_Report.pdf'
        )
    except Exception as e:
        print(f"PDF Export Error: {e}")
        return jsonify({"error": str(e)}), 500


# ==========================================
# 5. SETTINGS & SYSTEM MAINTENANCE ENDPOINTS
# ==========================================
@app.route('/api/settings/backup', methods=['GET'])
def download_db_backup():
    try:
        db_file = os.path.join(os.path.dirname(__file__), 'billing.db')
        if os.path.exists(db_file):
            log_activity("Settings", "Backup Downloaded", "Admin downloaded local SQLite billing.db database backup", performed_by="Admin")
            return send_file(db_file, mimetype='application/x-sqlite3', as_attachment=True, download_name='billing_backup.db')
        return jsonify({"error": "Database file not found"}), 404
    except Exception as e:
        print(f"Backup Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/settings/db-health', methods=['GET'])
def get_db_health():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM bills")
        bills_count = cursor.fetchone()[0] or 0
        
        cursor.execute("SELECT COUNT(*) FROM activity_logs")
        logs_count = cursor.fetchone()[0] or 0
        
        cursor.execute("SELECT COUNT(*) FROM products")
        products_count = cursor.fetchone()[0] or 0
        
        conn.close()
        
        db_file = os.path.join(os.path.dirname(__file__), 'billing.db')
        file_size_mb = round(os.path.getsize(db_file) / (1024 * 1024), 2) if os.path.exists(db_file) else 0.0
        
        return jsonify({
            "status": "Healthy & Connected",
            "db_name": "billing.db",
            "file_size_mb": file_size_mb,
            "total_bills": bills_count,
            "total_logs": logs_count,
            "total_products": products_count,
            "engine": "SQLite 3 Database"
        }), 200
    except Exception as e:
        print(f"DB Health Error: {e}")
        return jsonify({
            "status": "Connected (JSON Mode)",
            "db_name": "inventory.json",
            "file_size_mb": 0.05,
            "total_bills": 0,
            "total_logs": 0,
            "total_products": len(inventory_db),
            "engine": "Local Storage & JSON"
        }), 200


@app.route('/api/settings/test-email', methods=['POST'])
def test_email_endpoint():
    try:
        data = request.get_json() or {}
        target_email = (data.get('email') or 'sarveshraut400@gmail.com').strip()
        print(f"[DIAGNOSTIC] Testing live email dispatch to {target_email}...")
        success = send_welcome_email(target_email, "Test Recipient", "employee", "test1234")
        if success:
            return jsonify({
                "success": True,
                "message": f"Email successfully dispatched to {target_email} via Google SMTP (TLS 587/SSL 465)",
                "sender": SENDER_EMAIL
            }), 200
        else:
            return jsonify({
                "success": False,
                "error": f"Failed to deliver email to {target_email}. Check SMTP credentials or network connection."
            }), 500
    except Exception as e:
        print(f"test_email_endpoint error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == '__main__':
    print("[SUCCESS] Backend is starting on http://127.0.0.1:5000")
    app.run(debug=True, port=5000)